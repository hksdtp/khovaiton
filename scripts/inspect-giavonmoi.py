#!/usr/bin/env python3
"""
Script để kiểm tra cấu trúc file giavonmoi.xlsx
"""

import openpyxl
from pathlib import Path

def inspect_excel():
    """Kiểm tra cấu trúc file Excel"""
    excel_path = Path("giavonmoi.xlsx")
    if not excel_path.exists():
        print("❌ Không tìm thấy file giavonmoi.xlsx")
        return
    
    print("🔍 KIỂM TRA CẤU TRÚC FILE GIAVONMOI.XLSX")
    print("="*50)
    
    workbook = openpyxl.load_workbook(excel_path)
    
    for sheet_name in workbook.sheetnames:
        print(f"\n📊 SHEET: {sheet_name}")
        print("-" * 30)
        
        sheet = workbook[sheet_name]
        
        # Kiểm tra 10 dòng đầu
        print("📋 10 dòng đầu tiên:")
        for row_num in range(1, min(11, sheet.max_row + 1)):
            row_data = []
            for col_num in range(1, min(11, sheet.max_column + 1)):  # Chỉ lấy 10 cột đầu
                cell = sheet.cell(row=row_num, column=col_num)
                value = cell.value if cell.value else ""
                row_data.append(str(value)[:20])  # Cắt ngắn để dễ đọc
            print(f"   Dòng {row_num}: {row_data}")
        
        print(f"\n📏 Kích thước: {sheet.max_row} dòng x {sheet.max_column} cột")
        
        # Tìm dòng có thể là header
        print("\n🔍 Tìm dòng header có thể:")
        for row_num in range(1, min(21, sheet.max_row + 1)):
            row_values = []
            for col_num in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_num, column=col_num)
                if cell.value:
                    row_values.append(str(cell.value))
            
            # Kiểm tra xem có phải header không
            if any(keyword in ' '.join(row_values).lower() for keyword in ['mã', 'ma', 'code', 'tên', 'ten', 'name']):
                print(f"   Dòng {row_num} (có thể là header): {row_values}")

def main():
    inspect_excel()

if __name__ == "__main__":
    main()
