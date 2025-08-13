#!/usr/bin/env python3
"""
Script đơn giản để so sánh dữ liệu dự án với file giavonmoi.xlsx
Không cần pandas, sử dụng openpyxl để đọc Excel
"""

import json
import csv
import os
from pathlib import Path

def install_openpyxl():
    """Cài đặt openpyxl nếu chưa có"""
    try:
        import openpyxl
        return True
    except ImportError:
        print("📦 Đang cài đặt openpyxl...")
        os.system("pip3 install openpyxl")
        try:
            import openpyxl
            return True
        except ImportError:
            print("❌ Không thể cài đặt openpyxl")
            return False

def read_excel_simple():
    """Đọc file Excel đơn giản"""
    if not install_openpyxl():
        return None

    try:
        import openpyxl

        excel_path = Path("giavonmoi.xlsx")
        if not excel_path.exists():
            print("❌ Không tìm thấy file giavonmoi.xlsx")
            return None

        print("📖 Đang đọc file giavonmoi.xlsx...")

        workbook = openpyxl.load_workbook(excel_path)
        sheet_names = workbook.sheetnames
        print(f"📋 Sheets có sẵn: {sheet_names}")

        # Ưu tiên sheet "GV" hoặc "Vải tầng 4"
        sheet_name = sheet_names[0]
        if "GV" in sheet_names:
            sheet_name = "GV"
        elif "Vải tầng 4" in sheet_names:
            sheet_name = "Vải tầng 4"

        sheet = workbook[sheet_name]
        print(f"📊 Đọc sheet: {sheet_name}")

        # Đọc dữ liệu
        data = []
        headers = []

        # Đọc header (dòng 2 cho sheet GV và Vải tầng 4)
        header_row = 2 if sheet_name in ["GV", "Vải tầng 4"] else 1

        for cell in sheet[header_row]:
            headers.append(cell.value if cell.value else "")

        print(f"📋 Headers (dòng {header_row}): {headers}")

        # Đọc dữ liệu từ dòng sau header
        start_row = header_row + 1
        for row in sheet.iter_rows(min_row=start_row, values_only=True):
            if any(cell for cell in row):  # Bỏ qua dòng trống
                row_data = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_data[headers[i]] = value
                data.append(row_data)

        print(f"✅ Đã đọc {len(data)} dòng dữ liệu")
        return {'data': data, 'headers': headers, 'sheet_name': sheet_name}

    except Exception as e:
        print(f"❌ Lỗi đọc file Excel: {e}")
        return None

def read_csv_simple():
    """Đọc file CSV hiện tại"""
    csv_path = Path("public/fabric_inventory_updated.csv")
    if not csv_path.exists():
        print("❌ Không tìm thấy file fabric_inventory_updated.csv")
        return None
    
    try:
        data = []
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames
            for row in reader:
                data.append(row)
        
        print(f"✅ Đã đọc {len(data)} dòng từ CSV")
        print(f"📋 CSV Headers: {headers}")
        return {'data': data, 'headers': headers}
        
    except Exception as e:
        print(f"❌ Lỗi đọc CSV: {e}")
        return None

def normalize_code(code):
    """Chuẩn hóa mã vải"""
    if not code:
        return ""
    return str(code).strip().upper()

def find_code_column(headers):
    """Tìm cột chứa mã vải"""
    for header in headers:
        if header and any(keyword in str(header).lower() for keyword in ['mã hàng', 'ma_hang', 'ma', 'code', 'mã']):
            return header
    return None

def compare_simple(excel_data, csv_data):
    """So sánh dữ liệu đơn giản"""
    print("\n" + "="*60)
    print("🔍 BẮT ĐẦU SO SÁNH DỮ LIỆU")
    print("="*60)
    
    # Tìm cột mã vải
    excel_code_col = find_code_column(excel_data['headers'])
    csv_code_col = find_code_column(csv_data['headers'])
    
    if not excel_code_col:
        print("❌ Không tìm thấy cột mã vải trong Excel")
        return
    
    if not csv_code_col:
        print("❌ Không tìm thấy cột mã vải trong CSV")
        return
    
    print(f"📋 Excel sử dụng cột: {excel_code_col}")
    print(f"📋 CSV sử dụng cột: {csv_code_col}")
    
    # Thu thập mã vải
    excel_codes = set()
    csv_codes = set()
    
    excel_dict = {}
    csv_dict = {}
    
    # Xử lý Excel
    for row in excel_data['data']:
        code = normalize_code(row.get(excel_code_col))
        if code:
            excel_codes.add(code)
            excel_dict[code] = row
    
    # Xử lý CSV
    for row in csv_data['data']:
        code = normalize_code(row.get(csv_code_col))
        if code:
            csv_codes.add(code)
            csv_dict[code] = row
    
    # So sánh
    missing_in_csv = excel_codes - csv_codes
    extra_in_csv = csv_codes - excel_codes
    common_codes = excel_codes & csv_codes
    
    print(f"\n📊 KẾT QUẢ SO SÁNH:")
    print(f"   📈 Excel: {len(excel_codes)} mã vải")
    print(f"   📈 CSV:   {len(csv_codes)} mã vải")
    print(f"   ✅ Chung: {len(common_codes)} mã vải")
    print(f"   ❌ Thiếu trong CSV: {len(missing_in_csv)} mã vải")
    print(f"   ➕ Thừa trong CSV: {len(extra_in_csv)} mã vải")
    
    # Hiển thị một số mã thiếu
    if missing_in_csv:
        print(f"\n❌ MỘT SỐ MÃ THIẾU TRONG CSV (hiển thị 10 đầu):")
        for i, code in enumerate(list(missing_in_csv)[:10]):
            excel_row = excel_dict[code]
            # Tìm tên sản phẩm
            name = ""
            for col in excel_data['headers']:
                if col and any(keyword in str(col).lower() for keyword in ['ten', 'name', 'tên']):
                    name = excel_row.get(col, "")
                    break
            print(f"   {i+1}. {code} - {name}")
    
    # Hiển thị một số mã thừa
    if extra_in_csv:
        print(f"\n➕ MỘT SỐ MÃ THỪA TRONG CSV (hiển thị 10 đầu):")
        for i, code in enumerate(list(extra_in_csv)[:10]):
            csv_row = csv_dict[code]
            name = csv_row.get('Ten_hang', csv_row.get('Name', ""))
            print(f"   {i+1}. {code} - {name}")
    
    # Tạo file báo cáo
    create_simple_report(excel_codes, csv_codes, missing_in_csv, extra_in_csv, common_codes, excel_dict, csv_dict)
    
    return {
        'excel_total': len(excel_codes),
        'csv_total': len(csv_codes),
        'missing_in_csv': missing_in_csv,
        'extra_in_csv': extra_in_csv,
        'common_codes': common_codes
    }

def create_simple_report(excel_codes, csv_codes, missing_in_csv, extra_in_csv, common_codes, excel_dict, csv_dict):
    """Tạo báo cáo đơn giản"""
    
    report_content = f"""# 📊 BÁO CÁO SO SÁNH DỮ LIỆU - GIAVONMOI.XLSX

## 📈 Tổng quan:
- **Excel (giavonmoi.xlsx):** {len(excel_codes)} mã vải
- **CSV hiện tại:** {len(csv_codes)} mã vải  
- **Mã vải chung:** {len(common_codes)} mã vải
- **Thiếu trong CSV:** {len(missing_in_csv)} mã vải
- **Thừa trong CSV:** {len(extra_in_csv)} mã vải

## ❌ Thiếu trong CSV ({len(missing_in_csv)} mã):
"""
    
    for code in list(missing_in_csv)[:50]:  # Hiển thị 50 đầu
        report_content += f"- {code}\n"
    
    if len(missing_in_csv) > 50:
        report_content += f"- ... và {len(missing_in_csv) - 50} mã khác\n"
    
    report_content += f"""
## ➕ Thừa trong CSV ({len(extra_in_csv)} mã):
"""
    
    for code in list(extra_in_csv)[:50]:  # Hiển thị 50 đầu
        report_content += f"- {code}\n"
    
    if len(extra_in_csv) > 50:
        report_content += f"- ... và {len(extra_in_csv) - 50} mã khác\n"
    
    report_content += f"""
## 💡 Khuyến nghị:

### 🔧 Hành động cần thực hiện:
1. **Thêm {len(missing_in_csv)} mã vải thiếu** từ Excel vào CSV
2. **Kiểm tra {len(extra_in_csv)} mã vải thừa** trong CSV
3. **Đồng bộ dữ liệu** để đảm bảo tính nhất quán

### 📊 Tỷ lệ khớp:
- **Độ khớp:** {len(common_codes)}/{len(excel_codes)} = {(len(common_codes)/len(excel_codes)*100):.1f}%
- **Dữ liệu thiếu:** {(len(missing_in_csv)/len(excel_codes)*100):.1f}%
- **Dữ liệu thừa:** {(len(extra_in_csv)/len(csv_codes)*100):.1f}%

---
Tạo bởi: compare-data-simple.py
Thời gian: {__import__('datetime').datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
"""
    
    # Lưu báo cáo
    with open('BAO_CAO_SO_SANH_GIAVONMOI.md', 'w', encoding='utf-8') as f:
        f.write(report_content)
    
    # Tạo CSV danh sách thiếu
    if missing_in_csv:
        with open('ma_vai_thieu_trong_csv.csv', 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Ma_vai', 'Ghi_chu'])
            for code in missing_in_csv:
                writer.writerow([code, 'Có trong Excel nhưng thiếu trong CSV'])
        print(f"💾 Đã tạo: ma_vai_thieu_trong_csv.csv ({len(missing_in_csv)} mã)")
    
    # Tạo CSV danh sách thừa
    if extra_in_csv:
        with open('ma_vai_thua_trong_csv.csv', 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Ma_vai', 'Ghi_chu'])
            for code in extra_in_csv:
                writer.writerow([code, 'Có trong CSV nhưng không có trong Excel'])
        print(f"💾 Đã tạo: ma_vai_thua_trong_csv.csv ({len(extra_in_csv)} mã)")
    
    print(f"💾 Đã tạo báo cáo: BAO_CAO_SO_SANH_GIAVONMOI.md")

def main():
    print("🔍 SO SÁNH DỮ LIỆU VỚI GIAVONMOI.XLSX")
    print("="*50)
    
    # Đọc Excel
    excel_data = read_excel_simple()
    if not excel_data:
        return
    
    # Đọc CSV
    csv_data = read_csv_simple()
    if not csv_data:
        return
    
    # So sánh
    results = compare_simple(excel_data, csv_data)
    
    print("\n🎉 HOÀN TẤT!")
    print("📁 Kiểm tra các file báo cáo đã được tạo")

if __name__ == "__main__":
    main()
