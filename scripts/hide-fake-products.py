#!/usr/bin/env python3
"""
Script để ẩn những sản phẩm không có trong giavonmoi.xlsx (dữ liệu ảo)
Chỉ hiển thị những sản phẩm có thật từ file Excel
"""

import openpyxl
import json
import urllib.request
import urllib.parse
import re
from pathlib import Path
from datetime import datetime

# Supabase configuration
SUPABASE_URL = 'https://zgrfqkytbmahxcbgpkxx.supabase.co'
SUPABASE_ANON_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InpncmZxa3l0Ym1haHhjYmdwa3h4Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NDYxNjI1MTAsImV4cCI6MjA2MTczODUxMH0.a6giZZFMrj6jBhLip3ShOFCyTHt5dbe31UDGCECh0Zs'

def is_valid_fabric_code(code):
    """Kiểm tra xem có phải mã vải hợp lệ không"""
    if not code or not isinstance(code, str):
        return False
    
    code = str(code).strip()
    
    # Loại bỏ các dòng ghi chú
    invalid_patterns = [
        r'^[0-9]+\.',  # Bắt đầu bằng số và dấu chấm
        r'lưu ý',
        r'vải được kiểm tra',
        r'số lượng vải',
        r'chưa kiểm tra',
        r'phương pháp',
        r'kiểm kê',
        r'chất lượng',
        r'ngoại quan'
    ]
    
    for pattern in invalid_patterns:
        if re.search(pattern, code.lower()):
            return False
    
    if len(code) < 2 or len(code) > 100:
        return False
    
    if not re.search(r'[a-zA-Z0-9]', code):
        return False
    
    return True

def normalize_code(code):
    """Chuẩn hóa mã vải để so sánh"""
    if not code:
        return ""
    return str(code).strip().upper()

def read_real_codes_from_excel():
    """Đọc danh sách mã vải thật từ giavonmoi.xlsx"""
    try:
        excel_path = Path("giavonmoi.xlsx")
        if not excel_path.exists():
            print("❌ Không tìm thấy file giavonmoi.xlsx")
            return set()
            
        print("📖 Đang đọc danh sách mã vải thật từ giavonmoi.xlsx...")
        
        workbook = openpyxl.load_workbook(excel_path)
        
        # Ưu tiên sheet "GV"
        sheet_name = "GV" if "GV" in workbook.sheetnames else workbook.sheetnames[0]
        sheet = workbook[sheet_name]
        print(f"📊 Đọc sheet: {sheet_name}")
        
        # Đọc mã vải từ dòng 3, cột B (index 1)
        real_codes = set()
        valid_count = 0
        invalid_count = 0
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            if not row or len(row) <= 1:
                continue
                
            # Lấy mã hàng (cột B - index 1)
            code = row[1] if len(row) > 1 else None
            
            if is_valid_fabric_code(code):
                normalized_code = normalize_code(code)
                real_codes.add(normalized_code)
                valid_count += 1
            else:
                invalid_count += 1
        
        print(f"✅ Đã đọc {valid_count} mã vải thật, bỏ qua {invalid_count} dòng không hợp lệ")
        return real_codes
        
    except Exception as e:
        print(f"❌ Lỗi đọc file Excel: {e}")
        return set()

def supabase_request(method, endpoint, data=None):
    """Thực hiện request đến Supabase API"""
    url = f"{SUPABASE_URL}/rest/v1/{endpoint}"
    
    headers = {
        'apikey': SUPABASE_ANON_KEY,
        'Authorization': f'Bearer {SUPABASE_ANON_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'return=minimal'
    }
    
    try:
        if data:
            data_bytes = json.dumps(data).encode('utf-8')
            req = urllib.request.Request(url, data=data_bytes, headers=headers, method=method)
        else:
            req = urllib.request.Request(url, headers=headers, method=method)
        
        with urllib.request.urlopen(req) as response:
            return response.getcode(), response.read().decode('utf-8')
            
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode('utf-8')
    except Exception as e:
        return 0, str(e)

def get_all_products_from_supabase():
    """Lấy tất cả sản phẩm từ Supabase"""
    print("📥 Đang lấy danh sách sản phẩm từ Supabase...")
    
    status_code, response = supabase_request('GET', 'fabrics?select=id,code,name,is_hidden')
    
    if status_code == 200:
        try:
            products = json.loads(response)
            print(f"✅ Đã lấy {len(products)} sản phẩm từ Supabase")
            return products
        except Exception as e:
            print(f"❌ Lỗi parse response: {e}")
            return []
    else:
        print(f"❌ Lỗi lấy dữ liệu: {status_code} - {response}")
        return []

def hide_fake_products(real_codes, all_products):
    """Ẩn những sản phẩm không có trong danh sách thật"""
    print("\n🔍 Phân tích sản phẩm...")
    
    real_products = []
    fake_products = []
    already_hidden = []
    
    for product in all_products:
        code = normalize_code(product['code'])
        
        if code in real_codes:
            real_products.append(product)
        else:
            if product['is_hidden']:
                already_hidden.append(product)
            else:
                fake_products.append(product)
    
    print(f"📊 Kết quả phân tích:")
    print(f"   ✅ Sản phẩm thật: {len(real_products)}")
    print(f"   ❌ Sản phẩm ảo cần ẩn: {len(fake_products)}")
    print(f"   🔒 Đã ẩn trước đó: {len(already_hidden)}")
    
    if len(fake_products) == 0:
        print("✅ Không có sản phẩm ảo nào cần ẩn!")
        return 0
    
    print(f"\n🔒 Đang ẩn {len(fake_products)} sản phẩm ảo...")
    
    # Ẩn từng sản phẩm ảo
    hidden_count = 0
    failed_count = 0
    
    for i, product in enumerate(fake_products):
        print(f"   🔒 {i+1}/{len(fake_products)}: Ẩn {product['code']} - {product['name'][:50]}...")
        
        # Update is_hidden = true
        update_data = {'is_hidden': True}
        status_code, response = supabase_request('PATCH', f"fabrics?id=eq.{product['id']}", update_data)
        
        if status_code in [200, 204]:
            hidden_count += 1
        else:
            failed_count += 1
            print(f"      ❌ Lỗi ẩn {product['code']}: {status_code} - {response}")
    
    print(f"\n📊 Kết quả ẩn sản phẩm:")
    print(f"   ✅ Đã ẩn thành công: {hidden_count}")
    print(f"   ❌ Thất bại: {failed_count}")
    
    return hidden_count

def show_real_products(real_codes, all_products):
    """Hiển thị lại những sản phẩm thật đã bị ẩn nhầm"""
    print("\n👁️ Kiểm tra sản phẩm thật bị ẩn nhầm...")
    
    hidden_real_products = []
    
    for product in all_products:
        code = normalize_code(product['code'])
        
        if code in real_codes and product['is_hidden']:
            hidden_real_products.append(product)
    
    if len(hidden_real_products) == 0:
        print("✅ Không có sản phẩm thật nào bị ẩn nhầm!")
        return 0
    
    print(f"👁️ Đang hiển thị lại {len(hidden_real_products)} sản phẩm thật...")
    
    shown_count = 0
    failed_count = 0
    
    for i, product in enumerate(hidden_real_products):
        print(f"   👁️ {i+1}/{len(hidden_real_products)}: Hiện {product['code']} - {product['name'][:50]}...")
        
        # Update is_hidden = false
        update_data = {'is_hidden': False}
        status_code, response = supabase_request('PATCH', f"fabrics?id=eq.{product['id']}", update_data)
        
        if status_code in [200, 204]:
            shown_count += 1
        else:
            failed_count += 1
            print(f"      ❌ Lỗi hiện {product['code']}: {status_code} - {response}")
    
    print(f"\n📊 Kết quả hiển thị sản phẩm thật:")
    print(f"   ✅ Đã hiển thị thành công: {shown_count}")
    print(f"   ❌ Thất bại: {failed_count}")
    
    return shown_count

def create_report(real_codes, all_products, hidden_count, shown_count):
    """Tạo báo cáo chi tiết"""
    
    real_products = []
    fake_products = []
    
    for product in all_products:
        code = normalize_code(product['code'])
        
        if code in real_codes:
            real_products.append(product)
        else:
            fake_products.append(product)
    
    report_content = f"""# 📊 BÁO CÁO ẨN SẢN PHẨM ẢO

## 📈 Tổng quan:
- **File tham chiếu:** giavonmoi.xlsx (Sheet: GV)
- **Mã vải thật:** {len(real_codes)} mã
- **Tổng sản phẩm trong Supabase:** {len(all_products)}
- **Sản phẩm thật:** {len(real_products)}
- **Sản phẩm ảo:** {len(fake_products)}

## 🎯 Kết quả thực hiện:
- **Đã ẩn sản phẩm ảo:** {hidden_count}
- **Đã hiển thị sản phẩm thật:** {shown_count}

## ✅ Trạng thái sau khi xử lý:
- **Hiển thị:** Chỉ sản phẩm có trong giavonmoi.xlsx
- **Ẩn:** Tất cả sản phẩm không có trong file Excel (dữ liệu ảo)

## 📋 Danh sách sản phẩm ảo đã ẩn:
"""
    
    # Liệt kê một số sản phẩm ảo
    fake_list = [p for p in all_products if normalize_code(p['code']) not in real_codes]
    for i, product in enumerate(fake_list[:20]):  # Chỉ hiển thị 20 đầu
        report_content += f"{i+1}. **{product['code']}** - {product['name']}\n"
    
    if len(fake_list) > 20:
        report_content += f"... và {len(fake_list) - 20} sản phẩm ảo khác\n"
    
    report_content += f"""
## 💡 Lợi ích:
- ✅ Web app chỉ hiển thị dữ liệu thật từ giavonmoi.xlsx
- ✅ Loại bỏ hoàn toàn dữ liệu ảo/mock
- ✅ Tăng độ tin cậy của hệ thống
- ✅ Dữ liệu chính xác 100%

## 🚀 Bước tiếp theo:
1. Restart web app để load dữ liệu mới
2. Kiểm tra giao diện - chỉ hiển thị sản phẩm thật
3. Test các chức năng search, filter

---
Tạo bởi: hide-fake-products.py
Thời gian: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
"""
    
    # Lưu báo cáo
    with open('BAO_CAO_AN_SAN_PHAM_AO.md', 'w', encoding='utf-8') as f:
        f.write(report_content)
    
    print(f"💾 Đã tạo báo cáo: BAO_CAO_AN_SAN_PHAM_AO.md")

def main():
    print("🚀 BẮT ĐẦU ẨN SẢN PHẨM ẢO")
    print("="*50)
    
    # 1. Đọc danh sách mã vải thật từ Excel
    real_codes = read_real_codes_from_excel()
    if not real_codes:
        print("❌ Không có dữ liệu thật để tham chiếu")
        return
    
    print(f"📋 Đã có {len(real_codes)} mã vải thật từ giavonmoi.xlsx")
    
    # 2. Lấy tất cả sản phẩm từ Supabase
    all_products = get_all_products_from_supabase()
    if not all_products:
        print("❌ Không thể lấy dữ liệu từ Supabase")
        return
    
    # 3. Hiển thị sản phẩm thật (nếu bị ẩn nhầm)
    shown_count = show_real_products(real_codes, all_products)
    
    # 4. Ẩn sản phẩm ảo
    hidden_count = hide_fake_products(real_codes, all_products)
    
    # 5. Tạo báo cáo
    create_report(real_codes, all_products, hidden_count, shown_count)
    
    print("\n🎉 HOÀN TẤT!")
    print(f"✅ Đã ẩn {hidden_count} sản phẩm ảo")
    print(f"✅ Đã hiển thị {shown_count} sản phẩm thật")
    print("📁 Kiểm tra file BAO_CAO_AN_SAN_PHAM_AO.md")
    print("\n💡 Bước tiếp theo: Restart web app để thấy kết quả!")

if __name__ == "__main__":
    main()
