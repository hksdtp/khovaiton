#!/usr/bin/env python3
"""
Comprehensive Excel Formula Synchronization Script
Fixes all issues where Excel formulas weren't properly calculated in the database.
"""

import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import requests
import json
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime

# Supabase configuration
SUPABASE_URL = "https://zgrfqkytbmahxcbgpkxx.supabase.co"
SUPABASE_SERVICE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InpncmZxa3l0Ym1haHhjYmdwa3h4Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc0NjE2MjUxMCwiZXhwIjoyMDYxNzM4NTEwfQ.J2dd7QC-FozMWkDr33l1c9d6shWKefBpI7lxSL4v5JI"

# Excel file path
EXCEL_FILE = "/Users/ninh/Webapp/khovaiton/giavonmoi.xlsx"

def get_excel_formulas():
    """Extract all formulas from the Excel file and calculate their values."""
    print("🔍 Scanning Excel file for formulas...")
    
    # Load both versions - formulas and calculated values
    wb_formulas = load_workbook(EXCEL_FILE, data_only=False)
    wb_values = load_workbook(EXCEL_FILE, data_only=True)
    ws_formulas = wb_formulas.active
    ws_values = wb_values.active
    
    formulas = []
    column_names = {
        1: 'STT', 2: 'Mã hàng', 3: 'Tên hàng', 4: 'ĐVT', 5: 'Số lượng',
        6: 'Vị trí', 7: 'Loại Vải', 8: 'Tính trạng', 9: 'Giá vốn',
        10: 'Giá thanh lý', 11: 'Ghi chú', 12: 'Giá vải', 13: 'ĐV giá'
    }
    
    # Process all rows
    total_rows = ws_formulas.max_row
    
    for row_num in range(3, total_rows + 1):  # Start from row 3 (data rows)
        fabric_code = ws_formulas.cell(row=row_num, column=2).value
        if not fabric_code:
            continue
            
        # Check all columns for formulas
        for col_num in range(1, 14):
            cell = ws_formulas.cell(row=row_num, column=col_num)
            
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                calculated_value = ws_values.cell(row=row_num, column=col_num).value
                
                # Round to 2 decimal places for quantity fields
                if col_num == 5 and isinstance(calculated_value, (int, float)):
                    calculated_value = float(Decimal(str(calculated_value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                
                formulas.append({
                    'row': row_num,
                    'column': col_num,
                    'column_name': column_names.get(col_num, f'Col_{col_num}'),
                    'fabric_code': fabric_code,
                    'formula': cell.value,
                    'calculated_value': calculated_value
                })
    
    print(f"✅ Found {len(formulas)} formulas in Excel file")
    return formulas

def get_current_database_values(fabric_codes):
    """Get current values from the database for the specified fabric codes."""
    print("🔍 Fetching current database values...")
    
    headers = {
        'apikey': SUPABASE_SERVICE_KEY,
        'Authorization': f'Bearer {SUPABASE_SERVICE_KEY}',
        'Content-Type': 'application/json'
    }
    
    # Build query with all fabric codes
    codes_filter = ','.join([f'"{code}"' for code in fabric_codes])
    url = f"{SUPABASE_URL}/rest/v1/fabrics?code=in.({codes_filter})&select=code,quantity,price"
    
    response = requests.get(url, headers=headers)
    
    if response.status_code == 200:
        data = response.json()
        print(f"✅ Retrieved data for {len(data)} fabrics from database")
        return {item['code']: item for item in data}
    else:
        print(f"❌ Failed to fetch database data: {response.status_code}")
        print(response.text)
        return {}

def compare_values(formulas, db_values):
    """Compare Excel calculated values with database values."""
    print("🔍 Comparing Excel calculated values with database values...")
    
    discrepancies = []
    
    for formula in formulas:
        fabric_code = formula['fabric_code']
        excel_value = formula['calculated_value']
        
        if fabric_code in db_values:
            db_record = db_values[fabric_code]
            
            # Compare based on column
            if formula['column_name'] == 'Số lượng':
                db_value = db_record.get('quantity')
                if db_value != excel_value:
                    discrepancies.append({
                        'fabric_code': fabric_code,
                        'column': 'quantity',
                        'excel_value': excel_value,
                        'db_value': db_value,
                        'formula': formula['formula'],
                        'row': formula['row']
                    })
            elif formula['column_name'] in ['Giá vốn', 'Giá thanh lý', 'Giá vải']:
                db_value = db_record.get('price')
                if db_value != excel_value:
                    discrepancies.append({
                        'fabric_code': fabric_code,
                        'column': 'price',
                        'excel_value': excel_value,
                        'db_value': db_value,
                        'formula': formula['formula'],
                        'row': formula['row']
                    })
        else:
            print(f"⚠️  Fabric code '{fabric_code}' not found in database")
    
    print(f"🔍 Found {len(discrepancies)} discrepancies")
    return discrepancies

def generate_update_script(discrepancies):
    """Generate SQL update statements for all discrepancies."""
    if not discrepancies:
        print("✅ No discrepancies found - database is already synchronized!")
        return []
    
    print("📝 Generating update script...")
    
    updates = []
    for disc in discrepancies:
        update = {
            'fabric_code': disc['fabric_code'],
            'column': disc['column'],
            'new_value': disc['excel_value'],
            'old_value': disc['db_value'],
            'formula': disc['formula']
        }
        updates.append(update)
    
    return updates

def apply_database_updates(updates):
    """Apply the updates to the database."""
    if not updates:
        print("✅ No updates needed!")
        return True
    
    print(f"🚀 Applying {len(updates)} updates to database...")
    
    headers = {
        'apikey': SUPABASE_SERVICE_KEY,
        'Authorization': f'Bearer {SUPABASE_SERVICE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'return=minimal'
    }
    
    success_count = 0
    error_count = 0
    
    for update in updates:
        try:
            # Prepare update data
            update_data = {
                update['column']: update['new_value'],
                'updated_at': datetime.now().isoformat()
            }
            
            # Apply update
            url = f"{SUPABASE_URL}/rest/v1/fabrics?code=eq.{update['fabric_code']}"
            response = requests.patch(url, headers=headers, json=update_data)
            
            if response.status_code in [200, 204]:
                success_count += 1
                print(f"✅ Updated {update['fabric_code']}: {update['column']} = {update['new_value']}")
            else:
                error_count += 1
                print(f"❌ Failed to update {update['fabric_code']}: {response.status_code} - {response.text}")
                
        except Exception as e:
            error_count += 1
            print(f"❌ Error updating {update['fabric_code']}: {str(e)}")
    
    print(f"\n📊 Update Results:")
    print(f"   ✅ Successful: {success_count}")
    print(f"   ❌ Failed: {error_count}")
    
    return error_count == 0

def generate_summary_report(formulas, discrepancies, updates, success):
    """Generate a comprehensive summary report."""
    print("\n" + "="*80)
    print("📋 COMPREHENSIVE SYNCHRONIZATION REPORT")
    print("="*80)
    
    print(f"\n🔍 EXCEL ANALYSIS:")
    print(f"   • Total formulas found: {len(formulas)}")
    print(f"   • Formulas by column:")
    
    column_counts = {}
    for formula in formulas:
        col = formula['column_name']
        column_counts[col] = column_counts.get(col, 0) + 1
    
    for col, count in column_counts.items():
        print(f"     - {col}: {count}")
    
    print(f"\n📊 FORMULAS FOUND:")
    for i, formula in enumerate(formulas, 1):
        print(f"{i:2d}. Row {formula['row']:3d}: {formula['fabric_code']} - {formula['column_name']}")
        print(f"    Formula: {formula['formula']} = {formula['calculated_value']}")
    
    print(f"\n⚖️  COMPARISON RESULTS:")
    print(f"   • Total discrepancies: {len(discrepancies)}")
    
    if discrepancies:
        print(f"   • Discrepancies by fabric:")
        for disc in discrepancies:
            print(f"     - {disc['fabric_code']}: {disc['column']} ({disc['db_value']} → {disc['excel_value']})")
    
    print(f"\n🚀 UPDATE EXECUTION:")
    if updates:
        print(f"   • Updates applied: {len(updates)}")
        print(f"   • Success: {'✅ Yes' if success else '❌ No'}")
    else:
        print(f"   • No updates needed - data already synchronized!")
    
    print(f"\n🎯 FINAL STATUS:")
    if not discrepancies:
        print("   ✅ Database is fully synchronized with Excel formulas!")
    elif success:
        print("   ✅ All discrepancies have been resolved!")
    else:
        print("   ⚠️  Some updates failed - manual intervention may be required")
    
    print("\n" + "="*80)
    return True

def main():
    """Main execution function."""
    print("🚀 Starting Comprehensive Excel Formula Synchronization...")
    print("="*80)
    
    try:
        # Step 1: Get all formulas from Excel
        formulas = get_excel_formulas()
        
        if not formulas:
            print("✅ No formulas found in Excel file - nothing to synchronize!")
            return
        
        # Step 2: Get current database values
        fabric_codes = list(set(formula['fabric_code'] for formula in formulas))
        db_values = get_current_database_values(fabric_codes)
        
        # Step 3: Compare values
        discrepancies = compare_values(formulas, db_values)
        
        # Step 4: Generate update script
        updates = generate_update_script(discrepancies)
        
        # Step 5: Apply updates
        success = apply_database_updates(updates)
        
        # Step 6: Generate report
        generate_summary_report(formulas, discrepancies, updates, success)
        
        if success or not discrepancies:
            print("\n🎉 Synchronization completed successfully!")
        else:
            print("\n⚠️  Synchronization completed with some errors")
            
    except Exception as e:
        print(f"\n❌ Synchronization failed: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()