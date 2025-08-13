#!/usr/bin/env python3
"""
Script đơn giản để tạo SQL từ giavonmoi.xlsx
"""

import openpyxl
import json
import re
from pathlib import Path
from datetime import datetime

def is_valid_fabric_code(code):
    """Kiểm tra xem có phải mã vải hợp lệ không"""
    if not code or not isinstance(code, str):
        return False
    
    code = str(code).strip()
    
    # Loại bỏ các dòng ghi chú
    invalid_patterns = [
        r'^[0-9]+\.',  # Bắt đầu bằng số và dấu chấm
        r'lưu ý',
        r'vải được kiểm tra',
        r'số lượng vải',
        r'chưa kiểm tra',
        r'phương pháp',
        r'kiểm kê',
        r'chất lượng',
        r'ngoại quan'
    ]
    
    for pattern in invalid_patterns:
        if re.search(pattern, code.lower()):
            return False
    
    if len(code) < 2 or len(code) > 100:
        return False
    
    if not re.search(r'[a-zA-Z0-9]', code):
        return False
    
    return True

def normalize_quantity(qty_str):
    """Chuẩn hóa số lượng"""
    if not qty_str:
        return 0
    
    try:
        qty_str = str(qty_str).strip()
        return float(qty_str)
    except:
        return 0

def normalize_status(status_str):
    """Chuẩn hóa trạng thái"""
    if not status_str:
        return 'available'
    
    status = str(status_str).lower().strip()
    
    if any(keyword in status for keyword in ['lỗi', 'bẩn', 'mốc', 'hỏng']):
        return 'damaged'
    elif any(keyword in status for keyword in ['hết', 'không có']):
        return 'out_of_stock'
    elif any(keyword in status for keyword in ['ít', 'sắp hết']):
        return 'low_stock'
    else:
        return 'available'

def escape_sql_string(s):
    """Escape string cho SQL"""
    if not s:
        return "''"
    return "'" + str(s).replace("'", "''") + "'"

def read_real_data_from_excel():
    """Đọc dữ liệu thật từ giavonmoi.xlsx"""
    try:
        excel_path = Path("giavonmoi.xlsx")
        if not excel_path.exists():
            print("❌ Không tìm thấy file giavonmoi.xlsx")
            return []
            
        print("📖 Đang đọc dữ liệu thật từ giavonmoi.xlsx...")
        
        workbook = openpyxl.load_workbook(excel_path)
        
        # Ưu tiên sheet "GV"
        sheet_name = "GV" if "GV" in workbook.sheetnames else workbook.sheetnames[0]
        sheet = workbook[sheet_name]
        print(f"📊 Đọc sheet: {sheet_name}")
        
        # Đọc dữ liệu từ dòng 3
        fabrics = []
        valid_count = 0
        invalid_count = 0
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            if not row or len(row) == 0:
                continue
                
            # Lấy mã hàng (cột B - index 1)
            code = row[1] if len(row) > 1 else None
            
            if not is_valid_fabric_code(code):
                invalid_count += 1
                continue
            
            # Tạo fabric object
            fabric = {
                'code': str(code).strip(),
                'name': str(row[2]).strip() if len(row) > 2 and row[2] else f"Vải {code}",
                'unit': str(row[3]).strip() if len(row) > 3 and row[3] else 'm',
                'quantity': normalize_quantity(row[4] if len(row) > 4 else 0),
                'location': str(row[5]).strip() if len(row) > 5 and row[5] else 'Unknown',
                'type': str(row[6]).strip() if len(row) > 6 and row[6] else 'fabric',
                'status': normalize_status(row[7] if len(row) > 7 else ''),
                'note': str(row[10]).strip() if len(row) > 10 and row[10] else '',
            }
            
            fabrics.append(fabric)
            valid_count += 1
        
        print(f"✅ Đã đọc {valid_count} sản phẩm hợp lệ, bỏ qua {invalid_count} dòng không hợp lệ")
        return fabrics
        
    except Exception as e:
        print(f"❌ Lỗi đọc file Excel: {e}")
        return []

def generate_sql_script(fabrics):
    """Tạo SQL script để import vào Supabase"""
    
    sql_content = f"""-- SQL Script để import dữ liệu thật từ giavonmoi.xlsx
-- Tạo bởi: import-real-data-simple.py
-- Thời gian: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
-- Tổng số records: {len(fabrics)}

-- Xóa tất cả dữ liệu cũ
DELETE FROM fabrics;

-- Reset sequence (nếu cần)
ALTER SEQUENCE fabrics_id_seq RESTART WITH 1;

-- Insert dữ liệu thật
"""
    
    # Tạo INSERT statements
    for i, fabric in enumerate(fabrics):
        sql_content += f"""
INSERT INTO fabrics (
    code, name, type, quantity, unit, location, status, 
    image, price, price_note, is_hidden, note,
    created_at, updated_at
) VALUES (
    {escape_sql_string(fabric['code'])},
    {escape_sql_string(fabric['name'])},
    {escape_sql_string(fabric['type'])},
    {fabric['quantity']},
    {escape_sql_string(fabric['unit'])},
    {escape_sql_string(fabric['location'])},
    {escape_sql_string(fabric['status'])},
    '',  -- image sẽ được cập nhật sau
    NULL,  -- price
    NULL,  -- price_note
    false,  -- is_hidden
    {escape_sql_string(fabric['note'])},
    NOW(),
    NOW()
);"""
        
        # Thêm comment mỗi 50 records
        if (i + 1) % 50 == 0:
            sql_content += f"\n-- Đã insert {i + 1}/{len(fabrics)} records\n"
    
    sql_content += f"""

-- Kiểm tra kết quả
SELECT COUNT(*) as total_records FROM fabrics;
SELECT code, name, quantity, location FROM fabrics LIMIT 10;

-- Thống kê theo trạng thái
SELECT status, COUNT(*) as count FROM fabrics GROUP BY status;

-- Thống kê theo vị trí
SELECT location, COUNT(*) as count FROM fabrics GROUP BY location ORDER BY count DESC LIMIT 10;

-- Hoàn tất import {len(fabrics)} sản phẩm thật từ giavonmoi.xlsx
"""
    
    return sql_content

def generate_json_backup(fabrics):
    """Tạo file JSON backup"""
    backup_data = {
        'metadata': {
            'source': 'giavonmoi.xlsx',
            'sheet': 'GV',
            'total_records': len(fabrics),
            'created_at': datetime.now().isoformat(),
            'description': 'Dữ liệu thật từ file Excel giavonmoi.xlsx'
        },
        'fabrics': fabrics
    }
    
    return json.dumps(backup_data, ensure_ascii=False, indent=2)

def create_summary_report(fabrics):
    """Tạo báo cáo tổng hợp"""
    
    # Thống kê
    total_quantity = sum(f['quantity'] for f in fabrics)
    locations = set(f['location'] for f in fabrics)
    types = set(f['type'] for f in fabrics if f['type'])
    status_stats = {}
    for fabric in fabrics:
        status = fabric['status']
        status_stats[status] = status_stats.get(status, 0) + 1
    
    report_content = f"""# 📊 BÁO CÁO CHUẨN BỊ IMPORT DỮ LIỆU THẬT

## 📈 Tổng quan:
- **File nguồn:** giavonmoi.xlsx (Sheet: GV)
- **Dữ liệu đã chuẩn bị:** {len(fabrics)} sản phẩm
- **Tổng số lượng:** {total_quantity:.1f} đơn vị
- **Số vị trí:** {len(locations)} vị trí khác nhau
- **Số loại vải:** {len(types)} loại khác nhau

## 📊 Thống kê trạng thái:
"""
    
    for status, count in status_stats.items():
        status_name = {
            'available': 'Có sẵn',
            'damaged': 'Hỏng/Lỗi', 
            'low_stock': 'Sắp hết',
            'out_of_stock': 'Hết hàng'
        }.get(status, status)
        percentage = (count / len(fabrics) * 100)
        report_content += f"- **{status_name}:** {count} sản phẩm ({percentage:.1f}%)\n"
    
    # Top 10 sản phẩm
    top_products = sorted(fabrics, key=lambda x: x['quantity'], reverse=True)[:10]
    report_content += f"""
## 🏷️ Top 10 sản phẩm theo số lượng:
"""
    for i, product in enumerate(top_products, 1):
        report_content += f"{i}. **{product['code']}** - {product['name']} ({product['quantity']} {product['unit']})\n"
    
    # Top vị trí
    location_stats = {}
    for fabric in fabrics:
        loc = fabric['location']
        location_stats[loc] = location_stats.get(loc, 0) + 1
    
    top_locations = sorted(location_stats.items(), key=lambda x: x[1], reverse=True)[:10]
    report_content += f"""
## 📍 Top 10 vị trí có nhiều sản phẩm:
"""
    for i, (location, count) in enumerate(top_locations, 1):
        report_content += f"{i}. **{location}:** {count} sản phẩm\n"
    
    report_content += f"""
## 📁 Files được tạo:
- `import-real-data.sql` - SQL script để chạy trong Supabase
- `real-data-backup.json` - File backup dữ liệu JSON
- `BAO_CAO_CHUAN_BI_IMPORT.md` - Báo cáo này

## 🚀 Cách sử dụng:
1. **Mở Supabase SQL Editor**
2. **Copy nội dung file `import-real-data.sql`**
3. **Paste và chạy script**
4. **Kiểm tra kết quả**
5. **Restart web app**

## ⚠️ Lưu ý:
- Script sẽ XÓA TẤT CẢ dữ liệu cũ
- Backup được lưu trong file JSON
- Sau khi import, restart web app để load dữ liệu mới

---
Tạo bởi: import-real-data-simple.py
Thời gian: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
"""
    
    return report_content

def main():
    print("🚀 CHUẨN BỊ IMPORT DỮ LIỆU THẬT TỪ GIAVONMOI.XLSX")
    print("="*60)
    
    # 1. Đọc dữ liệu từ Excel
    fabrics = read_real_data_from_excel()
    if not fabrics:
        print("❌ Không có dữ liệu để xử lý")
        return
    
    print(f"\n📊 Đã chuẩn bị {len(fabrics)} sản phẩm thật")
    
    # 2. Tạo SQL script
    print("📝 Tạo SQL script...")
    sql_content = generate_sql_script(fabrics)
    with open('import-real-data.sql', 'w', encoding='utf-8') as f:
        f.write(sql_content)
    print("✅ Đã tạo file: import-real-data.sql")
    
    # 3. Tạo JSON backup
    print("💾 Tạo JSON backup...")
    json_content = generate_json_backup(fabrics)
    with open('real-data-backup.json', 'w', encoding='utf-8') as f:
        f.write(json_content)
    print("✅ Đã tạo file: real-data-backup.json")
    
    # 4. Tạo báo cáo
    print("📋 Tạo báo cáo...")
    report_content = create_summary_report(fabrics)
    with open('BAO_CAO_CHUAN_BI_IMPORT.md', 'w', encoding='utf-8') as f:
        f.write(report_content)
    print("✅ Đã tạo file: BAO_CAO_CHUAN_BI_IMPORT.md")
    
    print("\n🎉 HOÀN TẤT CHUẨN BỊ!")
    print("📁 Files đã tạo:")
    print("   • import-real-data.sql - SQL script để import")
    print("   • real-data-backup.json - Backup dữ liệu")
    print("   • BAO_CAO_CHUAN_BI_IMPORT.md - Báo cáo chi tiết")
    
    print("\n🚀 Bước tiếp theo:")
    print("1. Mở Supabase SQL Editor")
    print("2. Copy nội dung file import-real-data.sql")
    print("3. Paste và chạy script")
    print("4. Restart web app")

if __name__ == "__main__":
    main()
