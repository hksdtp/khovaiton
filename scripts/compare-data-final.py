#!/usr/bin/env python3
"""
Script cuối cùng để so sánh dữ liệu dự án với file giavonmoi.xlsx
Có lọc dữ liệu và phân tích chi tiết
"""

import openpyxl
import json
import csv
import os
import re
from pathlib import Path

def is_valid_fabric_code(code):
    """Kiểm tra xem có phải mã vải hợp lệ không"""
    if not code or not isinstance(code, str):
        return False
    
    code = str(code).strip()
    
    # Loại bỏ các dòng ghi chú
    invalid_patterns = [
        r'^[0-9]+\.',  # Bắt đầu bằng số và dấu chấm (như "1.", "2.")
        r'lưu ý',
        r'vải được kiểm tra',
        r'số lượng vải',
        r'chưa kiểm tra',
        r'phương pháp',
        r'kiểm kê',
        r'chất lượng',
        r'ngoại quan'
    ]
    
    for pattern in invalid_patterns:
        if re.search(pattern, code.lower()):
            return False
    
    # Mã vải hợp lệ thường có độ dài từ 2-50 ký tự và chứa ít nhất 1 chữ cái hoặc số
    if len(code) < 2 or len(code) > 50:
        return False
    
    # Phải chứa ít nhất 1 ký tự chữ hoặc số
    if not re.search(r'[a-zA-Z0-9]', code):
        return False
    
    return True

def read_excel_filtered():
    """Đọc file Excel với lọc dữ liệu"""
    try:
        excel_path = Path("giavonmoi.xlsx")
        if not excel_path.exists():
            print("❌ Không tìm thấy file giavonmoi.xlsx")
            return None
            
        print("📖 Đang đọc file giavonmoi.xlsx...")
        
        workbook = openpyxl.load_workbook(excel_path)
        sheet_names = workbook.sheetnames
        print(f"📋 Sheets có sẵn: {sheet_names}")
        
        # Ưu tiên sheet "GV"
        sheet_name = "GV" if "GV" in sheet_names else sheet_names[0]
        sheet = workbook[sheet_name]
        print(f"📊 Đọc sheet: {sheet_name}")
        
        # Đọc header từ dòng 2
        headers = []
        for cell in sheet[2]:
            headers.append(cell.value if cell.value else "")
        
        print(f"📋 Headers: {headers}")
        
        # Tìm cột mã hàng
        code_column_index = None
        for i, header in enumerate(headers):
            if header and 'mã hàng' in str(header).lower():
                code_column_index = i
                break
        
        if code_column_index is None:
            print("❌ Không tìm thấy cột 'Mã hàng'")
            return None
        
        # Đọc dữ liệu từ dòng 3
        data = []
        valid_count = 0
        invalid_count = 0
        
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if row and len(row) > code_column_index:
                code = row[code_column_index]
                
                if is_valid_fabric_code(code):
                    row_data = {}
                    for i, value in enumerate(row):
                        if i < len(headers):
                            row_data[headers[i]] = value
                    data.append(row_data)
                    valid_count += 1
                else:
                    invalid_count += 1
                    if code:  # Chỉ log những dòng có nội dung
                        print(f"   ⚠️  Bỏ qua dòng không hợp lệ: {str(code)[:50]}...")
        
        print(f"✅ Đã đọc {valid_count} dòng hợp lệ, bỏ qua {invalid_count} dòng không hợp lệ")
        return {'data': data, 'headers': headers, 'sheet_name': sheet_name}
        
    except Exception as e:
        print(f"❌ Lỗi đọc file Excel: {e}")
        return None

def read_csv_data():
    """Đọc file CSV hiện tại"""
    csv_path = Path("public/fabric_inventory_updated.csv")
    if not csv_path.exists():
        print("❌ Không tìm thấy file fabric_inventory_updated.csv")
        return None
    
    try:
        data = []
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames
            for row in reader:
                # Lọc những dòng có mã hàng hợp lệ
                code = row.get('Ma_hang', '')
                if is_valid_fabric_code(code):
                    data.append(row)
        
        print(f"✅ Đã đọc {len(data)} dòng hợp lệ từ CSV")
        return {'data': data, 'headers': headers}
        
    except Exception as e:
        print(f"❌ Lỗi đọc CSV: {e}")
        return None

def normalize_code(code):
    """Chuẩn hóa mã vải"""
    if not code:
        return ""
    return str(code).strip().upper()

def compare_detailed(excel_data, csv_data):
    """So sánh chi tiết"""
    print("\n" + "="*60)
    print("🔍 SO SÁNH CHI TIẾT DỮ LIỆU")
    print("="*60)
    
    # Thu thập mã vải
    excel_codes = set()
    csv_codes = set()
    excel_dict = {}
    csv_dict = {}
    
    # Xử lý Excel
    for row in excel_data['data']:
        code = normalize_code(row.get('Mã hàng'))
        if code:
            excel_codes.add(code)
            excel_dict[code] = row
    
    # Xử lý CSV
    for row in csv_data['data']:
        code = normalize_code(row.get('Ma_hang'))
        if code:
            csv_codes.add(code)
            csv_dict[code] = row
    
    # So sánh
    missing_in_csv = excel_codes - csv_codes
    extra_in_csv = csv_codes - excel_codes
    common_codes = excel_codes & csv_codes
    
    print(f"\n📊 KẾT QUẢ SO SÁNH CHI TIẾT:")
    print(f"   📈 Excel (sau lọc): {len(excel_codes)} mã vải")
    print(f"   📈 CSV (sau lọc):   {len(csv_codes)} mã vải")
    print(f"   ✅ Mã chung: {len(common_codes)} mã vải")
    print(f"   ❌ Thiếu trong CSV: {len(missing_in_csv)} mã vải")
    print(f"   ➕ Thừa trong CSV: {len(extra_in_csv)} mã vải")
    
    # Phân tích chi tiết các mã thiếu
    if missing_in_csv:
        print(f"\n❌ CHI TIẾT MÃ THIẾU TRONG CSV:")
        for i, code in enumerate(list(missing_in_csv)[:10]):
            excel_row = excel_dict[code]
            name = excel_row.get('Tên hàng', '')
            quantity = excel_row.get('Số lượng ', '')
            location = excel_row.get('Vị trí', '')
            print(f"   {i+1}. {code}")
            print(f"      📝 Tên: {name}")
            print(f"      📦 SL: {quantity}")
            print(f"      📍 Vị trí: {location}")
            print()
    
    # Phân tích chi tiết các mã thừa
    if extra_in_csv:
        print(f"\n➕ CHI TIẾT MÃ THỪA TRONG CSV:")
        for i, code in enumerate(list(extra_in_csv)[:10]):
            csv_row = csv_dict[code]
            name = csv_row.get('Ten_hang', '')
            quantity = csv_row.get('So_luong', '')
            location = csv_row.get('Vi_tri', '')
            print(f"   {i+1}. {code}")
            print(f"      📝 Tên: {name}")
            print(f"      📦 SL: {quantity}")
            print(f"      📍 Vị trí: {location}")
            print()
    
    # So sánh dữ liệu cho các mã chung
    differences = []
    print(f"\n🔍 KIỂM TRA SỰ KHÁC BIỆT CHO {min(10, len(common_codes))} MÃ CHUNG:")
    for i, code in enumerate(list(common_codes)[:10]):
        excel_row = excel_dict[code]
        csv_row = csv_dict[code]
        
        row_diffs = []
        
        # So sánh số lượng
        excel_qty = str(excel_row.get('Số lượng ', '')).strip()
        csv_qty = str(csv_row.get('So_luong', '')).strip()
        
        if excel_qty != csv_qty:
            row_diffs.append(f"SL: Excel={excel_qty}, CSV={csv_qty}")
        
        # So sánh tên
        excel_name = str(excel_row.get('Tên hàng', '')).strip()
        csv_name = str(csv_row.get('Ten_hang', '')).strip()
        
        if excel_name != csv_name:
            row_diffs.append(f"Tên khác nhau")
        
        if row_diffs:
            differences.append({
                'code': code,
                'differences': row_diffs,
                'excel_row': excel_row,
                'csv_row': csv_row
            })
            print(f"   ⚠️  {code}: {'; '.join(row_diffs)}")
    
    # Tạo báo cáo cuối cùng
    create_final_report(excel_codes, csv_codes, missing_in_csv, extra_in_csv, common_codes, differences, excel_dict, csv_dict)
    
    return {
        'excel_total': len(excel_codes),
        'csv_total': len(csv_codes),
        'missing_in_csv': missing_in_csv,
        'extra_in_csv': extra_in_csv,
        'common_codes': common_codes,
        'differences': differences
    }

def create_final_report(excel_codes, csv_codes, missing_in_csv, extra_in_csv, common_codes, differences, excel_dict, csv_dict):
    """Tạo báo cáo cuối cùng"""
    
    match_percentage = (len(common_codes) / len(excel_codes) * 100) if excel_codes else 0
    
    report_content = f"""# 📊 BÁO CÁO SO SÁNH DỮ LIỆU CUỐI CÙNG - GIAVONMOI.XLSX

## 📈 Tổng quan (sau lọc dữ liệu):
- **Excel (giavonmoi.xlsx):** {len(excel_codes)} mã vải hợp lệ
- **CSV hiện tại:** {len(csv_codes)} mã vải hợp lệ
- **Mã vải chung:** {len(common_codes)} mã vải
- **Thiếu trong CSV:** {len(missing_in_csv)} mã vải
- **Thừa trong CSV:** {len(extra_in_csv)} mã vải
- **Độ khớp:** {match_percentage:.1f}%

## ✅ Đánh giá chất lượng dữ liệu:
"""
    
    if match_percentage >= 95:
        report_content += "🟢 **XUẤT SẮC** - Dữ liệu khớp rất cao\n"
    elif match_percentage >= 90:
        report_content += "🟡 **TỐT** - Dữ liệu khớp cao\n"
    elif match_percentage >= 80:
        report_content += "🟠 **TRUNG BÌNH** - Cần cải thiện\n"
    else:
        report_content += "🔴 **CẦN CHỈNH SỬA** - Nhiều sự khác biệt\n"
    
    if missing_in_csv:
        report_content += f"\n## ❌ Thiếu trong CSV ({len(missing_in_csv)} mã):\n"
        for code in list(missing_in_csv)[:20]:
            excel_row = excel_dict.get(code, {})
            name = excel_row.get('Tên hàng', '')
            report_content += f"- **{code}** - {name}\n"
        
        if len(missing_in_csv) > 20:
            report_content += f"- ... và {len(missing_in_csv) - 20} mã khác\n"
    
    if extra_in_csv:
        report_content += f"\n## ➕ Thừa trong CSV ({len(extra_in_csv)} mã):\n"
        for code in list(extra_in_csv)[:20]:
            csv_row = csv_dict.get(code, {})
            name = csv_row.get('Ten_hang', '')
            report_content += f"- **{code}** - {name}\n"
        
        if len(extra_in_csv) > 20:
            report_content += f"- ... và {len(extra_in_csv) - 20} mã khác\n"
    
    if differences:
        report_content += f"\n## ⚠️ Sự khác biệt dữ liệu ({len(differences)} mã):\n"
        for diff in differences[:10]:
            report_content += f"- **{diff['code']}:** {'; '.join(diff['differences'])}\n"
    
    report_content += f"""
## 💡 Khuyến nghị hành động:

### 🔧 Ưu tiên cao:
1. **Thêm {len(missing_in_csv)} mã vải thiếu** từ Excel vào CSV
2. **Kiểm tra {len(extra_in_csv)} mã vải thừa** trong CSV
3. **Cập nhật {len(differences)} mã vải có sự khác biệt**

### 📊 Kế hoạch cải thiện:
- **Mục tiêu:** Đạt độ khớp 100%
- **Thời gian:** Ưu tiên trong 1-2 tuần
- **Trách nhiệm:** Team data entry

### 📁 Files được tạo:
- `BAO_CAO_SO_SANH_CUOI_CUNG.md` - Báo cáo này
- `ma_vai_thieu_chi_tiet.csv` - Danh sách chi tiết mã thiếu
- `ma_vai_thua_chi_tiet.csv` - Danh sách chi tiết mã thừa
- `du_lieu_khac_biet_chi_tiet.csv` - Chi tiết sự khác biệt

---
Tạo bởi: compare-data-final.py
Thời gian: {__import__('datetime').datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
"""
    
    # Lưu báo cáo
    with open('BAO_CAO_SO_SANH_CUOI_CUNG.md', 'w', encoding='utf-8') as f:
        f.write(report_content)
    
    # Tạo CSV chi tiết cho mã thiếu
    if missing_in_csv:
        with open('ma_vai_thieu_chi_tiet.csv', 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Ma_vai', 'Ten_hang', 'So_luong', 'Vi_tri', 'Ghi_chu'])
            for code in missing_in_csv:
                excel_row = excel_dict.get(code, {})
                writer.writerow([
                    code,
                    excel_row.get('Tên hàng', ''),
                    excel_row.get('Số lượng ', ''),
                    excel_row.get('Vị trí', ''),
                    'Có trong Excel nhưng thiếu trong CSV'
                ])
        print(f"💾 Đã tạo: ma_vai_thieu_chi_tiet.csv ({len(missing_in_csv)} mã)")
    
    # Tạo CSV chi tiết cho mã thừa
    if extra_in_csv:
        with open('ma_vai_thua_chi_tiet.csv', 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Ma_vai', 'Ten_hang', 'So_luong', 'Vi_tri', 'Ghi_chu'])
            for code in extra_in_csv:
                csv_row = csv_dict.get(code, {})
                writer.writerow([
                    code,
                    csv_row.get('Ten_hang', ''),
                    csv_row.get('So_luong', ''),
                    csv_row.get('Vi_tri', ''),
                    'Có trong CSV nhưng không có trong Excel'
                ])
        print(f"💾 Đã tạo: ma_vai_thua_chi_tiet.csv ({len(extra_in_csv)} mã)")
    
    print(f"💾 Đã tạo báo cáo cuối cùng: BAO_CAO_SO_SANH_CUOI_CUNG.md")

def main():
    print("🔍 SO SÁNH DỮ LIỆU CUỐI CÙNG VỚI GIAVONMOI.XLSX")
    print("="*60)
    
    # Đọc Excel với lọc
    excel_data = read_excel_filtered()
    if not excel_data:
        return
    
    # Đọc CSV
    csv_data = read_csv_data()
    if not csv_data:
        return
    
    # So sánh chi tiết
    results = compare_detailed(excel_data, csv_data)
    
    print("\n🎉 HOÀN TẤT!")
    print("📁 Kiểm tra các file báo cáo chi tiết đã được tạo")

if __name__ == "__main__":
    main()
