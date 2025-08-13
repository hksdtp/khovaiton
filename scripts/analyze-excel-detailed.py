#!/usr/bin/env python3
"""
Script để phân tích chi tiết file giavonmoi.xlsx
Tìm hiểu tại sao có sự khác biệt giữa số lượng đếm được và số lượng import thành công
"""

import openpyxl
import re
from pathlib import Path

def is_valid_fabric_code(code):
    """Kiểm tra xem có phải mã vải hợp lệ không"""
    if not code or not isinstance(code, str):
        return False
    
    code = str(code).strip()
    
    # Loại bỏ các dòng ghi chú
    invalid_patterns = [
        r'^[0-9]+\.',  # Bắt đầu bằng số và dấu chấm
        r'lưu ý',
        r'vải được kiểm tra',
        r'số lượng vải',
        r'chưa kiểm tra',
        r'phương pháp',
        r'kiểm kê',
        r'chất lượng',
        r'ngoại quan'
    ]
    
    for pattern in invalid_patterns:
        if re.search(pattern, code.lower()):
            return False
    
    if len(code) < 2 or len(code) > 100:
        return False
    
    if not re.search(r'[a-zA-Z0-9]', code):
        return False
    
    return True

def analyze_excel_detailed():
    """Phân tích chi tiết file Excel"""
    try:
        excel_path = Path("giavonmoi.xlsx")
        if not excel_path.exists():
            print("❌ Không tìm thấy file giavonmoi.xlsx")
            return
            
        print("🔍 PHÂN TÍCH CHI TIẾT FILE GIAVONMOI.XLSX")
        print("="*60)
        
        workbook = openpyxl.load_workbook(excel_path)
        
        # Liệt kê tất cả sheets
        print(f"📊 Sheets có sẵn: {workbook.sheetnames}")
        
        # Ưu tiên sheet "GV"
        sheet_name = "GV" if "GV" in workbook.sheetnames else workbook.sheetnames[0]
        sheet = workbook[sheet_name]
        print(f"📋 Đang phân tích sheet: {sheet_name}")
        
        # Đọc header từ dòng 2
        headers = []
        for cell in sheet[2]:
            headers.append(cell.value if cell.value else "")
        
        print(f"📝 Headers (dòng 2): {headers}")
        
        # Phân tích từng dòng
        all_rows = []
        valid_rows = []
        invalid_rows = []
        empty_rows = []
        duplicate_codes = {}
        
        total_rows = 0
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            total_rows += 1
            
            if not row or len(row) == 0 or all(cell is None or str(cell).strip() == '' for cell in row):
                empty_rows.append({
                    'row_num': row_num,
                    'content': 'Dòng trống'
                })
                continue
                
            # Lấy mã hàng (cột B - index 1)
            code = row[1] if len(row) > 1 else None
            
            row_info = {
                'row_num': row_num,
                'code': str(code).strip() if code else '',
                'name': str(row[2]).strip() if len(row) > 2 and row[2] else '',
                'full_row': [str(cell) if cell is not None else '' for cell in row[:11]]
            }
            
            all_rows.append(row_info)
            
            if is_valid_fabric_code(code):
                valid_rows.append(row_info)
                
                # Kiểm tra duplicate
                code_normalized = str(code).strip().upper()
                if code_normalized in duplicate_codes:
                    duplicate_codes[code_normalized].append(row_num)
                else:
                    duplicate_codes[code_normalized] = [row_num]
            else:
                invalid_rows.append(row_info)
        
        # Tìm duplicates
        actual_duplicates = {code: rows for code, rows in duplicate_codes.items() if len(rows) > 1}
        
        print(f"\n📊 THỐNG KÊ TỔNG QUAN:")
        print(f"   📋 Tổng số dòng từ dòng 3: {total_rows}")
        print(f"   📦 Dòng có dữ liệu: {len(all_rows)}")
        print(f"   ✅ Dòng hợp lệ: {len(valid_rows)}")
        print(f"   ❌ Dòng không hợp lệ: {len(invalid_rows)}")
        print(f"   🔳 Dòng trống: {len(empty_rows)}")
        print(f"   🔄 Mã trùng lặp: {len(actual_duplicates)} mã")
        
        # Chi tiết dòng không hợp lệ
        if invalid_rows:
            print(f"\n❌ CHI TIẾT DÒNG KHÔNG HỢP LỆ ({len(invalid_rows)} dòng):")
            for i, row in enumerate(invalid_rows[:10]):  # Chỉ hiển thị 10 dòng đầu
                print(f"   {i+1}. Dòng {row['row_num']}: '{row['code']}' - {row['name'][:50]}...")
            if len(invalid_rows) > 10:
                print(f"   ... và {len(invalid_rows) - 10} dòng khác")
        
        # Chi tiết dòng trống
        if empty_rows:
            print(f"\n🔳 CHI TIẾT DÒNG TRỐNG ({len(empty_rows)} dòng):")
            for i, row in enumerate(empty_rows[:5]):  # Chỉ hiển thị 5 dòng đầu
                print(f"   {i+1}. Dòng {row['row_num']}: {row['content']}")
            if len(empty_rows) > 5:
                print(f"   ... và {len(empty_rows) - 5} dòng khác")
        
        # Chi tiết mã trùng lặp
        if actual_duplicates:
            print(f"\n🔄 CHI TIẾT MÃ TRÙNG LẶP ({len(actual_duplicates)} mã):")
            for i, (code, rows) in enumerate(list(actual_duplicates.items())[:5]):
                print(f"   {i+1}. Mã '{code}' xuất hiện tại dòng: {rows}")
            if len(actual_duplicates) > 5:
                print(f"   ... và {len(actual_duplicates) - 5} mã khác")
        
        # Phân tích nguyên nhân khác biệt
        print(f"\n🔍 PHÂN TÍCH NGUYÊN NHÂN KHÁC BIỆT:")
        print(f"   📊 Bạn đếm được: 332 sản phẩm")
        print(f"   ✅ Script đọc được: {len(valid_rows)} sản phẩm hợp lệ")
        print(f"   📤 Đã import thành công: 325 sản phẩm")
        
        difference_count_vs_valid = 332 - len(valid_rows)
        difference_valid_vs_imported = len(valid_rows) - 325
        
        print(f"\n📈 PHÂN TÍCH SỰ KHÁC BIỆT:")
        print(f"   🔢 Khác biệt giữa đếm tay (332) và script ({len(valid_rows)}): {difference_count_vs_valid}")
        print(f"   🔢 Khác biệt giữa script ({len(valid_rows)}) và import (325): {difference_valid_vs_imported}")
        
        # Giải thích có thể
        print(f"\n💡 NGUYÊN NHÂN CÓ THỂ:")
        if difference_count_vs_valid > 0:
            print(f"   1. Có {difference_count_vs_valid} dòng bạn đếm nhưng script coi là không hợp lệ")
            print(f"      - Có thể là dòng ghi chú, header, hoặc dữ liệu không đúng format")
        
        if difference_valid_vs_imported > 0:
            print(f"   2. Có {difference_valid_vs_imported} sản phẩm hợp lệ nhưng không import được")
            print(f"      - Có thể do mã trùng lặp trong database")
            print(f"      - Hoặc lỗi khi insert vào Supabase")
        
        if len(actual_duplicates) > 0:
            print(f"   3. Có {len(actual_duplicates)} mã bị trùng lặp trong Excel")
            print(f"      - Database chỉ cho phép 1 record cho mỗi mã")
        
        # Tạo báo cáo chi tiết
        report_content = f"""# 🔍 BÁO CÁO PHÂN TÍCH CHI TIẾT GIAVONMOI.XLSX

## 📊 Thống kê tổng quan:
- **Tổng số dòng từ dòng 3:** {total_rows}
- **Dòng có dữ liệu:** {len(all_rows)}
- **Dòng hợp lệ:** {len(valid_rows)}
- **Dòng không hợp lệ:** {len(invalid_rows)}
- **Dòng trống:** {len(empty_rows)}
- **Mã trùng lặp:** {len(actual_duplicates)} mã

## 🔢 So sánh số liệu:
- **Bạn đếm được:** 332 sản phẩm
- **Script đọc được:** {len(valid_rows)} sản phẩm hợp lệ
- **Đã import thành công:** 325 sản phẩm

## 📈 Phân tích sự khác biệt:
- **Khác biệt đếm tay vs script:** {difference_count_vs_valid}
- **Khác biệt script vs import:** {difference_valid_vs_imported}

## ❌ Dòng không hợp lệ ({len(invalid_rows)} dòng):
"""
        
        for i, row in enumerate(invalid_rows[:20]):
            report_content += f"{i+1}. Dòng {row['row_num']}: '{row['code']}' - {row['name'][:50]}\n"
        
        if len(invalid_rows) > 20:
            report_content += f"... và {len(invalid_rows) - 20} dòng khác\n"
        
        if actual_duplicates:
            report_content += f"\n## 🔄 Mã trùng lặp ({len(actual_duplicates)} mã):\n"
            for i, (code, rows) in enumerate(list(actual_duplicates.items())[:10]):
                report_content += f"{i+1}. Mã '{code}' xuất hiện tại dòng: {rows}\n"
        
        report_content += f"""
## 💡 Kết luận:
Sự khác biệt giữa 332 (đếm tay) và 325 (import thành công) có thể do:
1. **{len(invalid_rows)} dòng không hợp lệ** (ghi chú, header, format sai)
2. **{len(actual_duplicates)} mã trùng lặp** (database chỉ lưu 1 record/mã)
3. **{len(empty_rows)} dòng trống**
4. **Lỗi import** một số records

## 🎯 Khuyến nghị:
- Kiểm tra lại các dòng không hợp lệ
- Xử lý mã trùng lặp trong Excel
- Đảm bảo format dữ liệu nhất quán

---
Tạo bởi: analyze-excel-detailed.py
"""
        
        # Lưu báo cáo
        with open('BAO_CAO_PHAN_TICH_EXCEL_CHI_TIET.md', 'w', encoding='utf-8') as f:
            f.write(report_content)
        
        print(f"\n💾 Đã tạo báo cáo chi tiết: BAO_CAO_PHAN_TICH_EXCEL_CHI_TIET.md")
        
        return {
            'total_rows': total_rows,
            'valid_rows': len(valid_rows),
            'invalid_rows': len(invalid_rows),
            'empty_rows': len(empty_rows),
            'duplicates': len(actual_duplicates),
            'all_data': all_rows,
            'valid_data': valid_rows,
            'invalid_data': invalid_rows
        }
        
    except Exception as e:
        print(f"❌ Lỗi phân tích file Excel: {e}")
        return None

def main():
    print("🔍 BẮT ĐẦU PHÂN TÍCH CHI TIẾT EXCEL")
    print("="*50)
    
    result = analyze_excel_detailed()
    
    if result:
        print(f"\n🎉 HOÀN TẤT PHÂN TÍCH!")
        print(f"📋 Kiểm tra file BAO_CAO_PHAN_TICH_EXCEL_CHI_TIET.md để xem chi tiết")

if __name__ == "__main__":
    main()
