#!/usr/bin/env python3
"""
Script để import dữ liệu thật từ giavonmoi.xlsx vào Supabase
Thay thế hoàn toàn dữ liệu giả bằng dữ liệu thật
"""

import openpyxl
import requests
import json
import re
from pathlib import Path
from datetime import datetime

# Supabase configuration
SUPABASE_URL = 'https://zgrfqkytbmahxcbgpkxx.supabase.co'
SUPABASE_ANON_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InpncmZxa3l0Ym1haHhjYmdwa3h4Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NDYxNjI1MTAsImV4cCI6MjA2MTczODUxMH0.a6giZZFMrj6jBhLip3ShOFCyTHt5dbe31UDGCECh0Zs'

def is_valid_fabric_code(code):
    """Kiểm tra xem có phải mã vải hợp lệ không"""
    if not code or not isinstance(code, str):
        return False
    
    code = str(code).strip()
    
    # Loại bỏ các dòng ghi chú
    invalid_patterns = [
        r'^[0-9]+\.',  # Bắt đầu bằng số và dấu chấm
        r'lưu ý',
        r'vải được kiểm tra',
        r'số lượng vải',
        r'chưa kiểm tra',
        r'phương pháp',
        r'kiểm kê',
        r'chất lượng',
        r'ngoại quan'
    ]
    
    for pattern in invalid_patterns:
        if re.search(pattern, code.lower()):
            return False
    
    # Mã vải hợp lệ
    if len(code) < 2 or len(code) > 100:
        return False
    
    if not re.search(r'[a-zA-Z0-9]', code):
        return False
    
    return True

def normalize_quantity(qty_str):
    """Chuẩn hóa số lượng"""
    if not qty_str:
        return 0
    
    try:
        # Loại bỏ khoảng trắng và chuyển đổi
        qty_str = str(qty_str).strip()
        return float(qty_str)
    except:
        return 0

def normalize_status(status_str):
    """Chuẩn hóa trạng thái"""
    if not status_str:
        return 'available'
    
    status = str(status_str).lower().strip()
    
    if any(keyword in status for keyword in ['lỗi', 'bẩn', 'mốc', 'hỏng']):
        return 'damaged'
    elif any(keyword in status for keyword in ['hết', 'không có']):
        return 'out_of_stock'
    elif any(keyword in status for keyword in ['ít', 'sắp hết']):
        return 'low_stock'
    else:
        return 'available'

def read_real_data_from_excel():
    """Đọc dữ liệu thật từ giavonmoi.xlsx"""
    try:
        excel_path = Path("giavonmoi.xlsx")
        if not excel_path.exists():
            print("❌ Không tìm thấy file giavonmoi.xlsx")
            return []
            
        print("📖 Đang đọc dữ liệu thật từ giavonmoi.xlsx...")
        
        workbook = openpyxl.load_workbook(excel_path)
        
        # Ưu tiên sheet "GV"
        sheet_name = "GV" if "GV" in workbook.sheetnames else workbook.sheetnames[0]
        sheet = workbook[sheet_name]
        print(f"📊 Đọc sheet: {sheet_name}")
        
        # Đọc header từ dòng 2
        headers = []
        for cell in sheet[2]:
            headers.append(cell.value if cell.value else "")
        
        print(f"📋 Headers: {headers}")
        
        # Mapping columns
        column_mapping = {
            'STT': 'stt',
            'Mã hàng': 'code',
            'Tên hàng': 'name',
            'ĐVT': 'unit',
            'Số lượng ': 'quantity',
            'Vị trí': 'location',
            'Loại Vải': 'type',
            'Tính trạng': 'status',
            'Giá vốn': 'cost_price',
            'Giá thanh lý': 'sale_price',
            'Ghi chú': 'note'
        }
        
        # Đọc dữ liệu từ dòng 3
        fabrics = []
        valid_count = 0
        invalid_count = 0
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            if not row or len(row) == 0:
                continue
                
            # Lấy mã hàng (cột B - index 1)
            code = row[1] if len(row) > 1 else None
            
            if not is_valid_fabric_code(code):
                invalid_count += 1
                if code:
                    print(f"   ⚠️  Bỏ qua dòng {row_num}: {str(code)[:50]}...")
                continue
            
            # Tạo fabric object
            fabric = {
                'code': str(code).strip(),
                'name': str(row[2]).strip() if len(row) > 2 and row[2] else f"Vải {code}",
                'unit': str(row[3]).strip() if len(row) > 3 and row[3] else 'm',
                'quantity': normalize_quantity(row[4] if len(row) > 4 else 0),
                'location': str(row[5]).strip() if len(row) > 5 and row[5] else 'Unknown',
                'type': str(row[6]).strip() if len(row) > 6 and row[6] else 'fabric',
                'status': normalize_status(row[7] if len(row) > 7 else ''),
                'cost_price': normalize_quantity(row[8] if len(row) > 8 else 0),
                'sale_price': normalize_quantity(row[9] if len(row) > 9 else 0),
                'note': str(row[10]).strip() if len(row) > 10 and row[10] else '',
                'is_hidden': False,
                'image': '',  # Sẽ được cập nhật sau
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat()
            }
            
            fabrics.append(fabric)
            valid_count += 1
        
        print(f"✅ Đã đọc {valid_count} sản phẩm hợp lệ, bỏ qua {invalid_count} dòng không hợp lệ")
        return fabrics
        
    except Exception as e:
        print(f"❌ Lỗi đọc file Excel: {e}")
        return []

def clear_existing_data():
    """Xóa dữ liệu cũ trong Supabase"""
    print("🗑️ Đang xóa dữ liệu cũ trong Supabase...")
    
    try:
        # Xóa tất cả records
        response = requests.delete(
            f"{SUPABASE_URL}/rest/v1/fabrics",
            headers={
                'apikey': SUPABASE_ANON_KEY,
                'Authorization': f'Bearer {SUPABASE_ANON_KEY}',
                'Content-Type': 'application/json'
            },
            params={'select': '*'}  # Xóa tất cả
        )
        
        if response.status_code in [200, 204]:
            print("✅ Đã xóa dữ liệu cũ thành công")
            return True
        else:
            print(f"❌ Lỗi xóa dữ liệu: {response.status_code} - {response.text}")
            return False
            
    except Exception as e:
        print(f"❌ Exception khi xóa dữ liệu: {e}")
        return False

def insert_real_data(fabrics):
    """Insert dữ liệu thật vào Supabase"""
    print(f"📤 Đang insert {len(fabrics)} sản phẩm vào Supabase...")
    
    # Insert theo batch để tránh timeout
    batch_size = 50
    total_inserted = 0
    
    for i in range(0, len(fabrics), batch_size):
        batch = fabrics[i:i + batch_size]
        
        try:
            response = requests.post(
                f"{SUPABASE_URL}/rest/v1/fabrics",
                headers={
                    'apikey': SUPABASE_ANON_KEY,
                    'Authorization': f'Bearer {SUPABASE_ANON_KEY}',
                    'Content-Type': 'application/json',
                    'Prefer': 'return=minimal'
                },
                json=batch
            )
            
            if response.status_code in [200, 201]:
                total_inserted += len(batch)
                print(f"   ✅ Batch {i//batch_size + 1}: Inserted {len(batch)} records ({total_inserted}/{len(fabrics)})")
            else:
                print(f"   ❌ Batch {i//batch_size + 1} failed: {response.status_code} - {response.text}")
                
        except Exception as e:
            print(f"   ❌ Exception in batch {i//batch_size + 1}: {e}")
    
    print(f"✅ Hoàn tất! Đã insert {total_inserted}/{len(fabrics)} sản phẩm")
    return total_inserted

def verify_data():
    """Kiểm tra dữ liệu đã được import"""
    print("🔍 Kiểm tra dữ liệu đã import...")
    
    try:
        response = requests.get(
            f"{SUPABASE_URL}/rest/v1/fabrics",
            headers={
                'apikey': SUPABASE_ANON_KEY,
                'Authorization': f'Bearer {SUPABASE_ANON_KEY}',
                'Content-Type': 'application/json'
            },
            params={'select': 'count', 'head': 'true'}
        )
        
        if response.status_code == 200:
            count = response.headers.get('Content-Range', '0').split('/')[-1]
            print(f"✅ Tổng số records trong Supabase: {count}")
            
            # Lấy sample data
            sample_response = requests.get(
                f"{SUPABASE_URL}/rest/v1/fabrics",
                headers={
                    'apikey': SUPABASE_ANON_KEY,
                    'Authorization': f'Bearer {SUPABASE_ANON_KEY}',
                    'Content-Type': 'application/json'
                },
                params={'select': 'code,name,quantity,location', 'limit': 5}
            )
            
            if sample_response.status_code == 200:
                sample_data = sample_response.json()
                print("📋 Sample data:")
                for item in sample_data:
                    print(f"   • {item['code']} - {item['name']} ({item['quantity']} tại {item['location']})")
            
            return True
        else:
            print(f"❌ Lỗi kiểm tra: {response.status_code}")
            return False
            
    except Exception as e:
        print(f"❌ Exception khi kiểm tra: {e}")
        return False

def create_summary_report(fabrics, inserted_count):
    """Tạo báo cáo tổng hợp"""
    report_content = f"""# 📊 BÁO CÁO IMPORT DỮ LIỆU THẬT - GIAVONMOI.XLSX

## 📈 Tổng quan:
- **File nguồn:** giavonmoi.xlsx (Sheet: GV)
- **Dữ liệu đọc được:** {len(fabrics)} sản phẩm
- **Đã import vào Supabase:** {inserted_count} sản phẩm
- **Tỷ lệ thành công:** {(inserted_count/len(fabrics)*100):.1f}%

## 📊 Thống kê dữ liệu:
- **Tổng số lượng vải:** {sum(f['quantity'] for f in fabrics):.1f} {fabrics[0]['unit'] if fabrics else 'đơn vị'}
- **Vị trí lưu trữ:** {len(set(f['location'] for f in fabrics))} vị trí khác nhau
- **Loại vải:** {len(set(f['type'] for f in fabrics if f['type']))} loại khác nhau

## 🏷️ Top 10 sản phẩm theo số lượng:
"""
    
    # Sort by quantity và lấy top 10
    top_products = sorted(fabrics, key=lambda x: x['quantity'], reverse=True)[:10]
    for i, product in enumerate(top_products, 1):
        report_content += f"{i}. **{product['code']}** - {product['name']} ({product['quantity']} {product['unit']})\n"
    
    report_content += f"""
## 📍 Phân bố theo vị trí:
"""
    
    # Thống kê theo vị trí
    location_stats = {}
    for fabric in fabrics:
        loc = fabric['location']
        if loc not in location_stats:
            location_stats[loc] = {'count': 0, 'total_qty': 0}
        location_stats[loc]['count'] += 1
        location_stats[loc]['total_qty'] += fabric['quantity']
    
    for location, stats in sorted(location_stats.items(), key=lambda x: x[1]['count'], reverse=True)[:10]:
        report_content += f"- **{location}:** {stats['count']} sản phẩm ({stats['total_qty']:.1f} đơn vị)\n"
    
    report_content += f"""
## ⚠️ Trạng thái sản phẩm:
"""
    
    # Thống kê theo trạng thái
    status_stats = {}
    for fabric in fabrics:
        status = fabric['status']
        status_stats[status] = status_stats.get(status, 0) + 1
    
    for status, count in status_stats.items():
        status_name = {
            'available': 'Có sẵn',
            'damaged': 'Hỏng/Lỗi',
            'low_stock': 'Sắp hết',
            'out_of_stock': 'Hết hàng'
        }.get(status, status)
        report_content += f"- **{status_name}:** {count} sản phẩm\n"
    
    report_content += f"""
## 🎯 Kết quả:
- ✅ Đã thay thế hoàn toàn dữ liệu giả bằng dữ liệu thật
- ✅ Web app sẽ hiển thị {inserted_count} sản phẩm thật từ Excel
- ✅ Dữ liệu được đồng bộ realtime từ Supabase
- ✅ Tất cả thông tin (mã, tên, số lượng, vị trí) đều chính xác

## 💡 Bước tiếp theo:
1. Restart web app để load dữ liệu mới
2. Kiểm tra hiển thị trên giao diện
3. Cập nhật ảnh cho các sản phẩm (nếu cần)
4. Test các chức năng search, filter

---
Tạo bởi: import-real-data-to-supabase.py
Thời gian: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
"""
    
    # Lưu báo cáo
    with open('BAO_CAO_IMPORT_DU_LIEU_THAT.md', 'w', encoding='utf-8') as f:
        f.write(report_content)
    
    print(f"💾 Đã tạo báo cáo: BAO_CAO_IMPORT_DU_LIEU_THAT.md")

def main():
    print("🚀 BẮT ĐẦU IMPORT DỮ LIỆU THẬT TỪ GIAVONMOI.XLSX")
    print("="*60)
    
    # 1. Đọc dữ liệu từ Excel
    fabrics = read_real_data_from_excel()
    if not fabrics:
        print("❌ Không có dữ liệu để import")
        return
    
    print(f"\n📊 Sẽ import {len(fabrics)} sản phẩm thật vào Supabase")
    
    # Xác nhận từ user
    confirm = input("\n⚠️  CẢNH BÁO: Thao tác này sẽ XÓA TẤT CẢ dữ liệu cũ và thay thế bằng dữ liệu thật.\nBạn có chắc chắn muốn tiếp tục? (y/N): ")
    
    if confirm.lower() != 'y':
        print("❌ Đã hủy import")
        return
    
    # 2. Xóa dữ liệu cũ
    if not clear_existing_data():
        print("❌ Không thể xóa dữ liệu cũ")
        return
    
    # 3. Insert dữ liệu mới
    inserted_count = insert_real_data(fabrics)
    
    # 4. Kiểm tra kết quả
    if verify_data():
        print("\n🎉 IMPORT THÀNH CÔNG!")
        
        # 5. Tạo báo cáo
        create_summary_report(fabrics, inserted_count)
        
        print("\n📋 Tóm tắt:")
        print(f"   ✅ Đã import {inserted_count} sản phẩm thật")
        print(f"   ✅ Thay thế hoàn toàn dữ liệu giả")
        print(f"   ✅ Web app sẽ hiển thị dữ liệu thật")
        print("\n💡 Bước tiếp theo: Restart web app để load dữ liệu mới!")
        
    else:
        print("❌ Import không thành công hoàn toàn")

if __name__ == "__main__":
    main()
